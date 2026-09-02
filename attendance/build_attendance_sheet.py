# -*- coding: utf-8 -*-
"""
勤怠管理（休暇調整）シートを生成するスクリプト。
python3 build_attendance_sheet.py [出力先.xlsx]

構成:
  使い方      … 運用ルールと入力手順
  設定        … メンバー、勤務曜日、営業曜日、記号、最低出勤人数、祝日
  休暇申請    … 唯一の入力元（1休暇1行）
  シフト表    … 年月を切り替えると自動で組み上がる月間一覧 + 人数判定
  有給管理    … 承認済み申請から自動集計
"""
import sys
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.comments import Comment

OUT = sys.argv[1] if len(sys.argv) > 1 else "勤怠管理_休暇調整シート.xlsx"
FONT = "Meiryo"

# ---- palette (portal と同系色) -------------------------------------------
NAVY = "1C1C3A"
ORANGE = "E8630A"
CREAM = "FAF8F4"
SAND = "F2EDE4"
BORDER = "E0D9CE"
MUTED = "8A8078"
INPUT_FILL = PatternFill("solid", fgColor="FFF6DD")   # 入力セル
AUTO_FILL = PatternFill("solid", fgColor="F3F3F3")    # 自動計算
HEAD_FILL = PatternFill("solid", fgColor=NAVY)
SUB_FILL = PatternFill("solid", fgColor=SAND)
OFF_FILL = PatternFill("solid", fgColor="FADBD8")     # 休み
HALF_FILL = PatternFill("solid", fgColor="FCF3CF")    # 半休
REMOTE_FILL = PatternFill("solid", fgColor="D6EAF8")  # 在宅・外出
PEND_FILL = PatternFill("solid", fgColor="FDEBD0")    # 申請中
CLOSED_FILL = PatternFill("solid", fgColor="D9D9D9")  # 休業日
NG_FILL = PatternFill("solid", fgColor="E74C3C")
WARN_FILL = PatternFill("solid", fgColor="F39C12")
OK_FILL = PatternFill("solid", fgColor="D5F5E3")

thin = Side(style="thin", color=BORDER)
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)


def f(bold=False, color="000000", size=10, italic=False):
    return Font(name=FONT, bold=bold, color=color, size=size, italic=italic)


def head(ws, row, col, text, width=None):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    c.fill = HEAD_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BOX
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def title(ws, text, sub=None):
    ws["B1"] = text
    ws["B1"].font = Font(name=FONT, bold=True, size=16, color=NAVY)
    if sub:
        ws["B2"] = sub
        ws["B2"].font = f(color=MUTED, size=9)
    ws.column_dimensions["A"].width = 2
    ws.sheet_view.showGridLines = False


def cell(ws, ref, value, fill=None, bold=False, color="000000", align=None, fmt=None, border=True):
    c = ws[ref]
    c.value = value
    c.font = f(bold=bold, color=color)
    if fill:
        c.fill = fill
    if align:
        c.alignment = Alignment(horizontal=align, vertical="center")
    if fmt:
        c.number_format = fmt
    if border:
        c.border = BOX
    return c


wb = Workbook()
from openpyxl.workbook.defined_name import DefinedName

def add_name(name, ref):
    dn = DefinedName(name, attr_text=ref)
    try:
        wb.defined_names[name] = dn
    except TypeError:
        wb.defined_names.append(dn)

# =============================================================================
# 1. 使い方
# =============================================================================
ws = wb.active
ws.title = "使い方"
title(ws, "勤怠管理（休暇調整）シート  使い方",
      "正社員2名 + 派遣2名の少人数チーム向け。誰がいつ休むかを一箇所で見える化し、出勤人数が最低ラインを割らないように調整する。")
ws.column_dimensions["B"].width = 4
ws.column_dimensions["C"].width = 110

rows = [
    ("■ 基本の考え方", True),
    ("・入力するのは「休暇申請」シートだけ。1休暇につき1行（連休は日ごとに1行ずつ）。", False),
    ("・「シフト表」は年・月を切り替えるだけで自動で組み上がる。手入力はしない。", False),
    ("・「有給管理」は承認済みの申請から自動集計。付与日と付与日数だけ最初に入れる。", False),
    ("・黄色のセル = 入力する場所。灰色のセル = 自動計算（触らない）。", False),
    ("", False),
    ("■ 最初にやること（5分）", True),
    ("1. 「設定」シートでメンバー4名の氏名・区分（正社員／派遣）・勤務曜日（○×）を入れる。", False),
    ("2. 「設定」の最低出勤人数（初期値 2名）と、最低正社員出勤人数（初期値 1名）を確認する。", False),
    ("3. 「有給管理」に各人の付与日・付与日数・繰越を入れる（派遣の残数は派遣元管理なので不要）。", False),
    ("4. サンプルとして入っている申請4行（9/10に3人が重なって「要確認」が出る例）は削除してよい。", False),
    ("", False),
    ("■ 休みを取りたいとき（申請者）", True),
    ("1. 「休暇申請」に 申請日・氏名・休暇日・種別・理由 を1行追加する。状態は「申請中」。", False),
    ("2. 右側の「判定」列に「要確認」と出たら、その日は既に誰かが休む予定。日程をずらせないか先に相談する。", False),
    ("3. 「シフト表」でも申請中の休みは (有) のようにカッコ付きで表示される。", False),
    ("", False),
    ("■ 承認するとき（管理者）", True),
    ("1. 「シフト表」の「申請中を承認した場合の判定」行を見る。「要注意」なら承認すると最低人数を割る。", False),
    ("2. 問題なければ「休暇申請」の状態を「承認」に変え、承認者名を入れる。却下なら「却下」。", False),
    ("3. 派遣メンバーの休みは、承認と同時に派遣元（派遣会社）にも連絡する。派遣元にとっては勤怠実績＝請求根拠。", False),
    ("", False),
    ("■ 運用ルールの例（設定シートで文言を変えられる）", True),
    ("・申請は原則、休みたい日の1週間前まで。急な体調不良は当日連絡＋事後に「欠」または「有」で登録。", False),
    ("・同じ日に希望が重なったら「先に申請した人」ではなく「直近3か月の休み取得日数が少ない人」を優先すると不公平感が出にくい。", False),
    ("・派遣2名だけの出勤日は作らない（指揮命令者が不在になる）。最低正社員出勤人数=1 はそのための設定。", False),
    ("・正社員は年5日の有給取得義務あり。「有給管理」の「年5日義務」列が「あと○日」のままなら、管理者側から取得日を提案する。", False),
    ("", False),
    ("■ 記号の意味", True),
    ("有=有給 / 午前=午前半休 / 午後=午後半休 / 休=公休・希望休 / 欠=欠勤 / 特=特別休暇 / 代=代休 / 在=在宅勤務 / 外=外出・出張", False),
    ("空欄=通常出勤、－=その人の勤務曜日ではない日、灰色の列=会社の休業日（土日祝）", False),
    ("", False),
    ("■ 注意", True),
    ("・シフト表は同時に1か月しか表示しない。過去月を見たいときは年・月を戻すだけでよい（データは休暇申請に残っている）。", False),
    ("・休暇申請の日付は必ず日付形式（2026/9/10 のように入力）。文字として入ると集計に乗らない。", False),
    ("・「休暇申請」の自動列（灰色）は200行分だけ式が入っている。それを超える場合は上の行をコピーする。", False),
]
r = 4
for text, bold in rows:
    c = ws.cell(row=r, column=3, value=text)
    c.font = Font(name=FONT, bold=bold, size=11 if bold else 10, color=NAVY if bold else "000000")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    if bold:
        ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=ORANGE)
    r += 1

# =============================================================================
# 2. 設定
# =============================================================================
st = wb.create_sheet("設定")
title(st, "設定", "黄色のセルを編集する。メンバーは最大6名まで。")
for col, w in zip("BCDEFGHIJKL", [14, 10, 5, 5, 5, 5, 5, 5, 5, 24, 3]):
    st.column_dimensions[col].width = w
st.column_dimensions["M"].width = 3
st.column_dimensions["N"].width = 14
st.column_dimensions["O"].width = 22

# --- メンバー
st["B3"] = "メンバー（氏名・区分・勤務曜日）"
st["B3"].font = f(bold=True, size=11, color=NAVY)
head(st, 4, 2, "氏名")
head(st, 4, 3, "区分")
for i, d in enumerate("月火水木金土日"):
    head(st, 4, 4 + i, d)
head(st, 4, 11, "備考")
members = [
    ("正社員A", "正社員", "○○○○○××", "氏名を書き換える"),
    ("正社員B", "正社員", "○○○○○××", ""),
    ("派遣A", "派遣", "○○○○○××", "契約勤務日に合わせて○×を直す"),
    ("派遣B", "派遣", "○○○○○××", ""),
    ("", "", "", ""),
    ("", "", "", ""),
]
STAFF_ROWS = range(5, 11)  # 5..10
for i, (name, kind, days, memo) in enumerate(members):
    row = 5 + i
    cell(st, f"B{row}", name, INPUT_FILL)
    cell(st, f"C{row}", kind, INPUT_FILL, align="center")
    for j in range(7):
        cell(st, f"{get_column_letter(4 + j)}{row}", days[j] if days else "", INPUT_FILL, align="center")
    cell(st, f"K{row}", memo, INPUT_FILL, color=MUTED)
st["B5"].comment = Comment("実際の氏名に書き換える。休暇申請・シフト表・有給管理はここを参照する。", "template")

dv_kind = DataValidation(type="list", formula1='"正社員,派遣"', allow_blank=True)
dv_ox = DataValidation(type="list", formula1='"○,×"', allow_blank=True)
st.add_data_validation(dv_kind)
st.add_data_validation(dv_ox)
dv_kind.add("C5:C10")
dv_ox.add("D5:J10")

# --- ルール
st["B12"] = "ルール（人数の下限）"
st["B12"].font = f(bold=True, size=11, color=NAVY)
head(st, 13, 2, "項目")
head(st, 13, 3, "値")
st.merge_cells("B13:B13")
cell(st, "B14", "最低出勤人数（1日あたり）")
cell(st, "C14", 2, INPUT_FILL, align="center")
cell(st, "B15", "最低正社員出勤人数")
cell(st, "C15", 1, INPUT_FILL, align="center")
cell(st, "B16", "申請期限の目安")
cell(st, "C16", "休みたい日の1週間前まで", INPUT_FILL)
st.merge_cells("C16:K16")
st["C14"].comment = Comment("4名体制で2名を下限にすると、同じ日に休めるのは最大2名。派遣だけの出勤日を避けたいなら下の正社員下限も合わせて使う。", "template")

# --- 営業曜日
st["B18"] = "営業曜日（×の曜日は休業日として全員カウント対象外）"
st["B18"].font = f(bold=True, size=11, color=NAVY)
head(st, 19, 2, "曜日")
head(st, 19, 3, "営業")
for i, d in enumerate("月火水木金土日"):
    row = 20 + i
    cell(st, f"B{row}", d, align="center")
    cell(st, f"C{row}", "○" if i < 5 else "×", INPUT_FILL, align="center")
dv_ox.add("C20:C26")

# --- 記号
st["B28"] = "記号（休暇申請の種別）"
st["B28"].font = f(bold=True, size=11, color=NAVY)
head(st, 29, 2, "記号")
head(st, 29, 3, "意味")
head(st, 29, 4, "出勤係数")
st.merge_cells("D29:E29")
head(st, 29, 6, "備考")
st.merge_cells("F29:K29")
symbols = [
    ("有", "有給休暇", 0, "正社員は年5日取得義務の対象"),
    ("午前", "午前半休", 0.5, "0.5日の有給として集計"),
    ("午後", "午後半休", 0.5, "0.5日の有給として集計"),
    ("休", "公休・希望休", 0, "無給の休み・シフト調整による休み"),
    ("特", "特別休暇（慶弔など）", 0, ""),
    ("代", "代休・振替休日", 0, ""),
    ("欠", "欠勤", 0, "事後登録用"),
    ("在", "在宅勤務", 1, "出勤扱い。所在把握のために記録"),
    ("外", "外出・出張", 1, "出勤扱い。事務所に人がいない場合は別途注意"),
]
SYM_FIRST, SYM_LAST = 30, 30 + len(symbols) - 1
for i, (s, mean, coef, memo) in enumerate(symbols):
    row = 30 + i
    cell(st, f"B{row}", s, align="center", bold=True)
    cell(st, f"C{row}", mean)
    cell(st, f"D{row}", coef, INPUT_FILL, align="center", fmt="0.0")
    st.merge_cells(f"D{row}:E{row}")
    cell(st, f"F{row}", memo, color=MUTED)
    st.merge_cells(f"F{row}:K{row}")
st["D30"].comment = Comment("その日を何人分の出勤として数えるか。1=出勤、0=休み、0.5=半日。", "template")

# --- 祝日
st["N3"] = "祝日・会社休業日（追加可）"
st["N3"].font = f(bold=True, size=11, color=NAVY)
head(st, 4, 14, "日付")
head(st, 4, 15, "名称")
holidays = [
    (date(2026, 9, 21), "敬老の日"), (date(2026, 9, 22), "国民の休日"), (date(2026, 9, 23), "秋分の日"),
    (date(2026, 10, 12), "スポーツの日"), (date(2026, 11, 3), "文化の日"), (date(2026, 11, 23), "勤労感謝の日"),
    (date(2026, 12, 29), "年末休業（例）"), (date(2026, 12, 30), "年末休業（例）"), (date(2026, 12, 31), "年末休業（例）"),
    (date(2027, 1, 1), "元日"), (date(2027, 1, 2), "年始休業（例）"), (date(2027, 1, 3), "年始休業（例）"),
    (date(2027, 1, 11), "成人の日"), (date(2027, 2, 11), "建国記念の日"), (date(2027, 2, 23), "天皇誕生日"),
    (date(2027, 3, 22), "春分の日 振替休日"), (date(2027, 4, 29), "昭和の日"),
    (date(2027, 5, 3), "憲法記念日"), (date(2027, 5, 4), "みどりの日"), (date(2027, 5, 5), "こどもの日"),
    (date(2027, 7, 19), "海の日"), (date(2027, 8, 11), "山の日"), (date(2027, 9, 20), "敬老の日"),
    (date(2027, 9, 23), "秋分の日"), (date(2027, 10, 11), "スポーツの日"), (date(2027, 11, 3), "文化の日"),
    (date(2027, 11, 23), "勤労感謝の日"),
]
HOL_RANGE = "HOLIDAYS"
for i, (d, name) in enumerate(holidays):
    row = 5 + i
    cell(st, f"N{row}", d, INPUT_FILL, fmt="yyyy/m/d", align="center")
    cell(st, f"O{row}", name, INPUT_FILL)
for row in range(5 + len(holidays), 81):
    cell(st, f"N{row}", None, INPUT_FILL, fmt="yyyy/m/d")
    cell(st, f"O{row}", None, INPUT_FILL)
st["N5"].comment = Comment("2026年9月〜2027年の祝日を入力済み（内閣府の祝日一覧に基づく。2027年分は公表後に要確認）。年末年始などの会社休業日も追加できる。", "template")
st.freeze_panes = "B5"

# =============================================================================
# 3. 休暇申請
# =============================================================================
lg = wb.create_sheet("休暇申請")
add_name("STAFF_NAMES", "設定!$B$5:$B$10")
add_name("STAFF_DAYS", "設定!$D$5:$J$10")
add_name("BIZ_DAYS", "設定!$C$20:$C$26")
add_name("HOLIDAYS", "設定!$N$5:$N$80")
add_name("MIN_ALL", "設定!$C$14")
add_name("MIN_FT", "設定!$C$15")
add_name("SYM_TBL", f"設定!$B${SYM_FIRST}:$D${SYM_LAST}")
title(lg, "休暇申請（ここだけ入力する）", "1休暇につき1行。連休は日ごとに1行。状態を「承認」にするとシフト表・有給管理に反映される。")
LOG_FIRST, LOG_LAST = 5, 204
add_name("LOG_NAME", f"休暇申請!$C${LOG_FIRST}:$C${LOG_LAST}")
add_name("LOG_DATE", f"休暇申請!$D${LOG_FIRST}:$D${LOG_LAST}")
add_name("LOG_SYM", f"休暇申請!$E${LOG_FIRST}:$E${LOG_LAST}")
add_name("LOG_STATE", f"休暇申請!$G${LOG_FIRST}:$G${LOG_LAST}")
add_name("KEY_A", f"休暇申請!$L${LOG_FIRST}:$L${LOG_LAST}")
add_name("KEY_P", f"休暇申請!$M${LOG_FIRST}:$M${LOG_LAST}")
cols = [
    ("申請日", 11), ("氏名", 12), ("休暇日", 11), ("種別", 8), ("理由・備考", 28),
    ("状態", 9), ("承認者", 10), ("同日の他の休み", 10), ("出勤見込み", 9), ("判定", 12),
    ("キー(承認)", 18), ("キー(申請中)", 18),
]
for i, (name, w) in enumerate(cols):
    head(lg, 4, 2 + i, name, w)
lg.row_dimensions[4].height = 30
lg["B3"] = "黄色 = 入力      灰色 = 自動（触らない）"
lg["B3"].font = f(color=MUTED, size=9)

samples = [
    (date(2026, 9, 1), "正社員A", date(2026, 9, 10), "有", "私用", "承認", "正社員B"),
    (date(2026, 9, 2), "派遣A", date(2026, 9, 10), "有", "通院", "申請中", ""),
    (date(2026, 9, 2), "正社員B", date(2026, 9, 15), "午前", "役所手続き", "承認", "正社員A"),
    (date(2026, 9, 2), "派遣B", date(2026, 9, 10), "休", "私用（希望休）", "申請中", ""),
]
staff_names = "STAFF_NAMES"
staff_days = "STAFF_DAYS"
for row in range(LOG_FIRST, LOG_LAST + 1):
    s = samples[row - LOG_FIRST] if row - LOG_FIRST < len(samples) else None
    cell(lg, f"B{row}", s[0] if s else None, INPUT_FILL, fmt="yyyy/m/d", align="center")
    cell(lg, f"C{row}", s[1] if s else None, INPUT_FILL)
    cell(lg, f"D{row}", s[2] if s else None, INPUT_FILL, fmt="yyyy/m/d", align="center")
    cell(lg, f"E{row}", s[3] if s else None, INPUT_FILL, align="center")
    cell(lg, f"F{row}", s[4] if s else None, INPUT_FILL)
    cell(lg, f"G{row}", s[5] if s else None, INPUT_FILL, align="center")
    cell(lg, f"H{row}", s[6] if s else None, INPUT_FILL)
    # 同日の他の休み（承認 + 申請中、自分を除く）
    cell(lg, f"I{row}",
         f'=IF(D{row}="","",COUNTIFS(LOG_DATE,D{row},LOG_STATE,"承認")'
         f'+COUNTIFS(LOG_DATE,D{row},LOG_STATE,"申請中")-IF(G{row}="却下",0,1))',
         AUTO_FILL, align="center")
    # 出勤見込み = その曜日に勤務予定の人数 − 休む人数（承認+申請中、自分含む）
    cell(lg, f"J{row}",
         f'=IF(D{row}="","",IFERROR(SUMPRODUCT((INDEX(STAFF_DAYS,0,WEEKDAY(D{row},2))="○")*1)-I{row}-IF(G{row}="却下",0,1),"日付を確認"))',
         AUTO_FILL, align="center")
    cell(lg, f"K{row}",
         f'=IF(D{row}="","",IFERROR(IF(OR(INDEX(BIZ_DAYS,WEEKDAY(D{row},2))="×",COUNTIF(HOLIDAYS,D{row})>0),"休業日",IF(J{row}<MIN_ALL,"要確認","OK")),"日付を確認"))',
         AUTO_FILL, align="center", bold=True)
    cell(lg, f"L{row}", f'=IF(AND(G{row}="承認",C{row}<>"",D{row}<>""),C{row}&"|"&TEXT(D{row},"yyyymmdd"),"")', AUTO_FILL, color=MUTED)
    cell(lg, f"M{row}", f'=IF(AND(G{row}="申請中",C{row}<>"",D{row}<>""),C{row}&"|"&TEXT(D{row},"yyyymmdd"),"")', AUTO_FILL, color=MUTED)
lg["I5"].comment = Comment("同じ日に休む予定（承認済み＋申請中）の人数。自分は除く。半休も1人として数える（安全側）。", "template")
lg["J5"].comment = Comment("その曜日に勤務予定の人数 − 休む人数（自分を含む、承認＋申請中）。半休も1人分の休みとして数えるので、実際より厳しめに出る。", "template")
lg["K5"].comment = Comment("出勤見込みが設定の最低出勤人数を割ると「要確認」。承認前に日程調整を。", "template")

dv_name = DataValidation(type="list", formula1=f"={staff_names}", allow_blank=True)
dv_sym = DataValidation(type="list", formula1=f"=設定!$B${SYM_FIRST}:$B${SYM_LAST}", allow_blank=True)
dv_state = DataValidation(type="list", formula1='"申請中,承認,却下"', allow_blank=True)
dv_date = DataValidation(type="date", operator="greaterThan", formula1="40000", allow_blank=True,
                         error="日付を 2026/9/10 の形式で入力してください", showErrorMessage=True)
for dv in (dv_name, dv_sym, dv_state, dv_date):
    lg.add_data_validation(dv)
dv_name.add(f"C{LOG_FIRST}:C{LOG_LAST}")
dv_sym.add(f"E{LOG_FIRST}:E{LOG_LAST}")
dv_state.add(f"G{LOG_FIRST}:G{LOG_LAST}")
dv_date.add(f"B{LOG_FIRST}:B{LOG_LAST}")
dv_date.add(f"D{LOG_FIRST}:D{LOG_LAST}")

lg.conditional_formatting.add(f"K{LOG_FIRST}:K{LOG_LAST}", CellIsRule(operator="equal", formula=['"要確認"'], fill=WARN_FILL, font=Font(name=FONT, bold=True, color="FFFFFF")))
lg.conditional_formatting.add(f"K{LOG_FIRST}:K{LOG_LAST}", CellIsRule(operator="equal", formula=['"OK"'], fill=OK_FILL))
lg.conditional_formatting.add(f"G{LOG_FIRST}:G{LOG_LAST}", CellIsRule(operator="equal", formula=['"申請中"'], fill=PEND_FILL))
lg.conditional_formatting.add(f"G{LOG_FIRST}:G{LOG_LAST}", CellIsRule(operator="equal", formula=['"承認"'], fill=OK_FILL))
lg.conditional_formatting.add(f"G{LOG_FIRST}:G{LOG_LAST}", CellIsRule(operator="equal", formula=['"却下"'], fill=CLOSED_FILL))
lg.freeze_panes = "B5"
lg.auto_filter.ref = f"B4:M{LOG_LAST}"

# =============================================================================
# 4. シフト表
# =============================================================================
sh = wb.create_sheet("シフト表")
title(sh, "シフト表（自動）", "年と月を変えるだけ。休暇申請の内容が自動で入る。( ) 付きは申請中。")
sh.column_dimensions["B"].width = 12
sh.column_dimensions["C"].width = 8
DAY_FIRST_COL, DAY_LAST_COL = 4, 34  # D..AH
for c in range(DAY_FIRST_COL, DAY_LAST_COL + 1):
    sh.column_dimensions[get_column_letter(c)].width = 4.6
sh.column_dimensions["AI"].width = 2
for col, w in zip(["AJ", "AK", "AL", "AM"], [8, 8, 8, 8]):
    sh.column_dimensions[col].width = w

# 年月入力
cell(sh, "B3", "年", SUB_FILL, bold=True, align="center")
cell(sh, "C3", 2026, INPUT_FILL, bold=True, align="center", fmt="0")
cell(sh, "D3", "月", SUB_FILL, bold=True, align="center")
cell(sh, "E3", 9, INPUT_FILL, bold=True, align="center", fmt="0")
sh.merge_cells("E3:F3")
cell(sh, "H3", "最低出勤", SUB_FILL, bold=True, align="center")
sh.merge_cells("H3:J3")
cell(sh, "K3", "=設定!C14", AUTO_FILL, align="center", color="008000")
cell(sh, "L3", "最低正社員", SUB_FILL, bold=True, align="center")
sh.merge_cells("L3:O3")
cell(sh, "P3", "=設定!C15", AUTO_FILL, align="center", color="008000")
sh["C3"].comment = Comment("表示したい年。", "template")
sh["E3"].comment = Comment("表示したい月（1〜12）。", "template")

R_DATE, R_WD, R_CLOSED = 5, 6, 7
R_STAFF_FIRST = 8               # 8..13
R_STAFF_LAST = R_STAFF_FIRST + 5
R_CNT, R_CNT_FT, R_CNT_TMP, R_JUDGE, R_CNT_PEND, R_JUDGE_PEND = 15, 16, 17, 18, 19, 20
R_HELP_TITLE = 22
R_HA_FIRST = 23                 # 承認済み係数 23..28
R_HA_LAST = R_HA_FIRST + 5
R_HP_FIRST = 29                 # 申請中含む係数 29..34
R_HP_LAST = R_HP_FIRST + 5

head(sh, R_DATE, 2, "氏名")
head(sh, R_DATE, 3, "区分")
sh.merge_cells(start_row=R_DATE, start_column=2, end_row=R_CLOSED, end_column=2)
sh.merge_cells(start_row=R_DATE, start_column=3, end_row=R_CLOSED, end_column=3)
head(sh, R_DATE, 36, "所定日数")
head(sh, R_DATE, 37, "出勤日数")
head(sh, R_DATE, 38, "休暇日数")
head(sh, R_DATE, 39, "有給日数")
for col in (36, 37, 38, 39):
    sh.merge_cells(start_row=R_DATE, start_column=col, end_row=R_CLOSED, end_column=col)
sh.row_dimensions[R_DATE].height = 18

sym_tbl = "SYM_TBL"
log_sym = "LOG_SYM"
log_keyA = "KEY_A"
log_keyP = "KEY_P"

for c in range(DAY_FIRST_COL, DAY_LAST_COL + 1):
    L = get_column_letter(c)
    P = get_column_letter(c - 1)
    # 日付
    if c == DAY_FIRST_COL:
        dcell = cell(sh, f"{L}{R_DATE}", "=DATE($C$3,$E$3,1)", SUB_FILL, bold=True, align="center", fmt="d")
    else:
        dcell = cell(sh, f"{L}{R_DATE}", f'=IF({P}{R_DATE}="","",IF(DAY({P}{R_DATE}+1)=1,"",{P}{R_DATE}+1))',
                     SUB_FILL, bold=True, align="center", fmt="d")
    cell(sh, f"{L}{R_WD}", f'=IF({L}{R_DATE}="","",CHOOSE(WEEKDAY({L}{R_DATE},2),"月","火","水","木","金","土","日"))',
         SUB_FILL, align="center")
    cell(sh, f"{L}{R_CLOSED}",
         f'=IF({L}{R_DATE}="","",IF(OR(INDEX(BIZ_DAYS,WEEKDAY({L}{R_DATE},2))="×",COUNTIF({HOL_RANGE},{L}{R_DATE})>0),"休業",""))',
         SUB_FILL, align="center", color=MUTED)
    # メンバー行
    for i in range(6):
        r = R_STAFF_FIRST + i
        key = f'$B{r}&"|"&TEXT({L}${R_DATE},"yyyymmdd")'
        cell(sh, f"{L}{r}",
             f'=IF(OR($B{r}="",{L}${R_DATE}=""),"",IF({L}${R_CLOSED}="休業","",'
             f'IF(INDEX({staff_days},MATCH($B{r},{staff_names},0),WEEKDAY({L}${R_DATE},2))="×","－",'
             f'IFERROR(INDEX({log_sym},MATCH({key},{log_keyA},0)),'
             f'IFERROR("("&INDEX({log_sym},MATCH({key},{log_keyP},0))&")","")))))',
             None, align="center")
        # 係数（承認済み）
        ha = R_HA_FIRST + i
        cell(sh, f"{L}{ha}",
             f'=IF(OR($B{r}="",{L}${R_DATE}="",{L}${R_CLOSED}="休業",{L}{r}="－"),"",'
             f'IF({L}{r}="",1,IFERROR(VLOOKUP({L}{r},{sym_tbl},3,0),1)))',
             AUTO_FILL, align="center", color=MUTED, fmt="0.#")
        # 係数（申請中も休み扱い）
        hp = R_HP_FIRST + i
        cell(sh, f"{L}{hp}",
             f'=IF(OR($B{r}="",{L}${R_DATE}="",{L}${R_CLOSED}="休業",{L}{r}="－"),"",'
             f'IF({L}{r}="",1,IFERROR(VLOOKUP(SUBSTITUTE(SUBSTITUTE({L}{r},"(",""),")",""),{sym_tbl},3,0),1)))',
             AUTO_FILL, align="center", color=MUTED, fmt="0.#")
    # 集計
    HA = f"{L}{R_HA_FIRST}:{L}{R_HA_LAST}"
    HP = f"{L}{R_HP_FIRST}:{L}{R_HP_LAST}"
    KIND = f"$C${R_STAFF_FIRST}:$C${R_STAFF_LAST}"
    closed = f'{L}${R_DATE}="","",IF({L}${R_CLOSED}="休業","－",'
    cell(sh, f"{L}{R_CNT}", f'=IF({closed}SUM({HA})))', AUTO_FILL, bold=True, align="center", fmt="0.#")
    cell(sh, f"{L}{R_CNT_FT}", f'=IF({closed}SUMIF({KIND},"正社員",{HA})))', AUTO_FILL, align="center", fmt="0.#")
    cell(sh, f"{L}{R_CNT_TMP}", f'=IF({closed}SUMIF({KIND},"派遣",{HA})))', AUTO_FILL, align="center", fmt="0.#")
    cell(sh, f"{L}{R_JUDGE}",
         f'=IF({closed}IF(OR({L}{R_CNT}<MIN_ALL,{L}{R_CNT_FT}<MIN_FT),"NG","OK")))',
         AUTO_FILL, bold=True, align="center")
    cell(sh, f"{L}{R_CNT_PEND}", f'=IF({closed}SUM({HP})))', AUTO_FILL, align="center", fmt="0.#")
    cell(sh, f"{L}{R_JUDGE_PEND}",
         f'=IF({closed}IF(OR({L}{R_CNT_PEND}<MIN_ALL,SUMIF({KIND},"正社員",{HP})<MIN_FT),"要注意","OK")))',
         AUTO_FILL, bold=True, align="center")

# メンバー名・区分（設定から参照）と右側集計
for i in range(6):
    r = R_STAFF_FIRST + i
    cell(sh, f"B{r}", f"=設定!B{5 + i}", None, bold=True, color="008000")
    cell(sh, f"C{r}", f"=設定!C{5 + i}", None, color="008000", align="center")
    ha = R_HA_FIRST + i
    hp = R_HP_FIRST + i
    cell(sh, f"B{ha}", f"=B{r}", AUTO_FILL, color=MUTED)
    cell(sh, f"C{ha}", "承認済", AUTO_FILL, color=MUTED, align="center")
    cell(sh, f"B{hp}", f"=B{r}", AUTO_FILL, color=MUTED)
    cell(sh, f"C{hp}", "申請中込", AUTO_FILL, color=MUTED, align="center")
    row_days = f"D{r}:AH{r}"
    row_ha = f"D{ha}:AH{ha}"
    cell(sh, f"AJ{r}", f'=IF($B{r}="","",COUNT({row_ha}))', AUTO_FILL, align="center")
    cell(sh, f"AK{r}", f'=IF($B{r}="","",SUM({row_ha}))', AUTO_FILL, align="center", fmt="0.#")
    cell(sh, f"AL{r}", f'=IF($B{r}="","",AJ{r}-AK{r})', AUTO_FILL, align="center", fmt="0.#")
    cell(sh, f"AM{r}", f'=IF($B{r}="","",COUNTIF({row_days},"有")+0.5*(COUNTIF({row_days},"午前")+COUNTIF({row_days},"午後")))',
         AUTO_FILL, align="center", fmt="0.#")

labels = {
    R_CNT: "出勤人数（承認済み反映）",
    R_CNT_FT: "　うち正社員",
    R_CNT_TMP: "　うち派遣",
    R_JUDGE: "判定（NG=下限割れ）",
    R_CNT_PEND: "申請中も休みにした場合の出勤人数",
    R_JUDGE_PEND: "申請中を承認した場合の判定",
}
for r, text in labels.items():
    cell(sh, f"B{r}", text, SUB_FILL, bold=r in (R_CNT, R_JUDGE, R_JUDGE_PEND))
    sh.merge_cells(f"B{r}:C{r}")
cell(sh, f"B{R_HELP_TITLE}", "自動計算エリア（出勤係数。編集不要）", None, color=MUTED, border=False)
sh.merge_cells(f"B{R_HELP_TITLE}:L{R_HELP_TITLE}")

# 条件付き書式
GRID = f"D{R_STAFF_FIRST}:AH{R_STAFF_LAST}"
ALL = f"D{R_DATE}:AH{R_JUDGE_PEND}"
sh.conditional_formatting.add(ALL, FormulaRule(formula=[f'D${R_CLOSED}="休業"'], fill=CLOSED_FILL, stopIfTrue=True))
sh.conditional_formatting.add(f"D{R_WD}:AH{R_WD}", FormulaRule(formula=[f'D{R_WD}="日"'], font=Font(name=FONT, color="C0392B", bold=True)))
sh.conditional_formatting.add(f"D{R_WD}:AH{R_WD}", FormulaRule(formula=[f'D{R_WD}="土"'], font=Font(name=FONT, color="2E86C1", bold=True)))
sh.conditional_formatting.add(GRID, FormulaRule(formula=[f'LEFT(D{R_STAFF_FIRST},1)="("'], fill=PEND_FILL, stopIfTrue=True))
sh.conditional_formatting.add(GRID, FormulaRule(formula=[f'OR(D{R_STAFF_FIRST}="有",D{R_STAFF_FIRST}="休",D{R_STAFF_FIRST}="特",D{R_STAFF_FIRST}="代",D{R_STAFF_FIRST}="欠")'], fill=OFF_FILL, font=Font(name=FONT, bold=True, color="922B21")))
sh.conditional_formatting.add(GRID, FormulaRule(formula=[f'OR(D{R_STAFF_FIRST}="午前",D{R_STAFF_FIRST}="午後")'], fill=HALF_FILL))
sh.conditional_formatting.add(GRID, FormulaRule(formula=[f'OR(D{R_STAFF_FIRST}="在",D{R_STAFF_FIRST}="外")'], fill=REMOTE_FILL))
sh.conditional_formatting.add(GRID, FormulaRule(formula=[f'D{R_STAFF_FIRST}="－"'], font=Font(name=FONT, color="BBBBBB")))
sh.conditional_formatting.add(f"D{R_JUDGE}:AH{R_JUDGE}", CellIsRule(operator="equal", formula=['"NG"'], fill=NG_FILL, font=Font(name=FONT, bold=True, color="FFFFFF")))
sh.conditional_formatting.add(f"D{R_JUDGE}:AH{R_JUDGE}", CellIsRule(operator="equal", formula=['"OK"'], fill=OK_FILL))
sh.conditional_formatting.add(f"D{R_JUDGE_PEND}:AH{R_JUDGE_PEND}", CellIsRule(operator="equal", formula=['"要注意"'], fill=WARN_FILL, font=Font(name=FONT, bold=True, color="FFFFFF")))
sh.conditional_formatting.add(f"D{R_JUDGE_PEND}:AH{R_JUDGE_PEND}", CellIsRule(operator="equal", formula=['"OK"'], fill=OK_FILL))

# 凡例
legend = [("有/休/特/代/欠", OFF_FILL), ("午前/午後", HALF_FILL), ("在/外", REMOTE_FILL), ("(申請中)", PEND_FILL), ("休業日", CLOSED_FILL)]
col = 36
for text, fill in legend:
    c = sh.cell(row=3, column=col, value=text)
    c.fill = fill
    c.font = f(size=9)
    c.alignment = Alignment(horizontal="center")
    col += 1
sh.column_dimensions["AN"].width = 10
sh.freeze_panes = f"D{R_STAFF_FIRST}"
sh.sheet_properties.tabColor = ORANGE

# =============================================================================
# 5. 有給管理
# =============================================================================
pm = wb.create_sheet("有給管理")
title(pm, "有給管理（自動集計）", "承認済みの「有」「午前」「午後」を基準日以降で数える。派遣の残数は派遣元が管理するため、ここでは取得日数の把握のみ。")
heads = [("氏名", 14), ("区分", 9), ("基準日（付与日）", 14), ("付与日数", 9), ("繰越日数", 9),
         ("取得日数（承認済）", 12), ("残日数", 12), ("年5日取得義務", 14), ("備考", 30)]
for i, (name, w) in enumerate(heads):
    head(pm, 4, 2 + i, name, w)
pm.row_dimensions[4].height = 30
lgB = "LOG_NAME"
lgD = "LOG_DATE"
lgE = "LOG_SYM"
lgG = "LOG_STATE"
defaults = [(date(2026, 4, 1), 10, 0), (date(2026, 4, 1), 11, 3), (None, None, None), (None, None, None), (None, None, None), (None, None, None)]
for i in range(6):
    r = 5 + i
    cell(pm, f"B{r}", f"=設定!B{5 + i}", None, bold=True, color="008000")
    cell(pm, f"C{r}", f"=設定!C{5 + i}", None, color="008000", align="center")
    d, grant, carry = defaults[i]
    cell(pm, f"D{r}", d, INPUT_FILL, fmt="yyyy/m/d", align="center")
    cell(pm, f"E{r}", grant, INPUT_FILL, align="center", fmt="0.#")
    cell(pm, f"F{r}", carry, INPUT_FILL, align="center", fmt="0.#")
    since = f'">="&IF(D{r}="",0,D{r})'
    cnt = lambda s: f'COUNTIFS({lgB},B{r},{lgD},{since},{lgE},"{s}",{lgG},"承認")'
    cell(pm, f"G{r}", f'=IF(B{r}="","",{cnt("有")}+0.5*({cnt("午前")}+{cnt("午後")}))', AUTO_FILL, align="center", fmt="0.#")
    cell(pm, f"H{r}", f'=IF(B{r}="","",IF(C{r}="派遣","派遣元で管理",E{r}+F{r}-G{r}))', AUTO_FILL, align="center", fmt="0.#")
    cell(pm, f"I{r}", f'=IF(B{r}="","",IF(C{r}="派遣","対象外（派遣元）",IF(E{r}>=10,IF(G{r}>=5,"達成","あと"&(5-G{r})&"日"),"対象外")))',
         AUTO_FILL, align="center")
    cell(pm, f"J{r}", "" if i else "付与日数・繰越は例。実際の値に直す。", INPUT_FILL, color=MUTED)
pm["D5"].comment = Comment("直近の有給付与日。この日以降の承認済み取得を数える。", "template")
pm["E5"].comment = Comment("付与日数の例（入社6か月で10日、1年6か月で11日 …）。実際の値に直す。", "template")
pm["I5"].comment = Comment("年10日以上付与される人は、付与日から1年以内に5日取得させる義務がある（労基法39条7項）。", "template")
pm.conditional_formatting.add("I5:I10", FormulaRule(formula=['LEFT(I5,2)="あと"'], fill=WARN_FILL, font=Font(name=FONT, bold=True, color="FFFFFF")))
pm.conditional_formatting.add("I5:I10", CellIsRule(operator="equal", formula=['"達成"'], fill=OK_FILL))
pm.conditional_formatting.add("H5:H10", FormulaRule(formula=['AND(ISNUMBER(H5),H5<=2)'], fill=OFF_FILL))
pm.freeze_panes = "B5"

# sheet order: 使い方, 設定, 休暇申請, シフト表, 有給管理  → シフト表を先頭に見せたいので並び替え
wb._sheets = [wb["シフト表"], wb["休暇申請"], wb["有給管理"], wb["設定"], wb["使い方"]]
wb.active = 0
wb.save(OUT)
print("saved", OUT)
